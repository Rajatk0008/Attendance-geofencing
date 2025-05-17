import pandas as pd
from io import BytesIO
from flask_mail import Message
from datetime import datetime, time
from sqlalchemy import and_
import calendar
from openpyxl.styles import PatternFill

from core import mail, db
from core.models import Attendance, User

def generate_and_send_monthly_report(app, year, month):
    with app.app_context():
        start_date = datetime(year, month, 1).date()
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day).date()

        # Get all users first for full report coverage
        users = User.query.all()
        user_map = {u.id: u for u in users}

        # Prepare date range list
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Get all attendance records in the range
        records = Attendance.query.filter(
            and_(
                Attendance.date >= start_date,
                Attendance.date <= end_date
            )
        ).all()

        # Organize attendance by date and user
        attendance_dict = {}
        for rec in records:
            attendance_dict.setdefault(rec.date, {})[rec.user_id] = rec

        rows = []
        for date in date_range:
            for user in users:
                rec = attendance_dict.get(date, {}).get(user.id)
                punch_in = rec.punch_in_time.strftime('%H:%M') if rec and rec.punch_in_time else ''
                punch_out = rec.punch_out_time.strftime('%H:%M') if rec and rec.punch_out_time else ''
                status = 'Present' if punch_in or punch_out else 'Absent'

                rows.append({
                    'Name': user.name,
                    'Email': user.email,
                    'Date': date.strftime('%Y-%m-%d'),
                    'Punch In': punch_in,
                    'Punch Out': punch_out,
                    'Status': status,
                    'IP Address': rec.device_ip if rec else ''
                })
            # Insert empty row after each date
            rows.append({})  # empty dict will become a blank row

        df = pd.DataFrame(rows)

        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance')
            ws = writer.sheets['Attendance']

            # Style "Absent" rows red
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row in range(2, ws.max_row + 1):  # Skip header row
                cell = ws.cell(row=row, column=6)  # Status column is 6th column (A=1, F=6)
                if cell.value == 'Absent':
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = red_fill

        excel_file.seek(0)

        recipients = [u.email for u in User.query.filter(User.role.in_(['superadmin'])).all()]
        if not recipients:
            print("No superadmin users to email.")
            return

        msg = Message(
            subject=f"📊 Attendance Report - {start_date.strftime('%B %Y')}",
            recipients=recipients,
            body=f"Attached is the attendance report for {start_date.strftime('%B %Y')}."
        )
        msg.attach(
            f"attendance_{year}_{month}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            excel_file.read()
        )
        mail.send(msg)
        print(f"Sent monthly report for {month}/{year} to {', '.join(recipients)}.")
